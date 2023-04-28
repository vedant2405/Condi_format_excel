import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# read the Excel sheet into a Pandas dataframe
df = pd.read_excel('filename.xlsx')

# define a function to highlight rows based on project
def highlight_project(row):
    if row['Project'] in ['P1', 'P2', 'P3', 'P4', 'P5']:
        return [True]*len(row)
    else:
        return [False]*len(row)

# create a new dataframe with the highlighted rows
highlighted_df = df.style.apply(highlight_project, axis=1)\
                 .set_table_styles([{'selector': '', 'props': [('border', '1px solid black')]}])\
                 .background_gradient(subset=['Project'], cmap='YlOrBr')\
                 .set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})

# write the highlighted dataframe to the same Excel sheet
book = load_workbook('filename.xlsx')
writer = pd.ExcelWriter('filename.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for sheetname in writer.sheets:
    sheet = writer.sheets[sheetname]
    for idx, row in highlighted_df.iterrows():
        for colno, value in enumerate(row):
            cell = sheet.cell(row=idx+2, column=colno+1)
            cell.value = value
            if isinstance(value, pd._libs.tslibs.timestamps.Timestamp):
                cell.number_format = 'yyyy-mm-dd'
            elif isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            if row.isin(['P1', 'P2', 'P3', 'P4', 'P5']).any():
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                for cell in sheet[idx+2]:
                    cell.fill = fill
    writer.save()

print('Done!')
