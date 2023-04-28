import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# read the Excel sheet into a Pandas dataframe
df = pd.read_excel('filename.xlsx')

# define a function to highlight rows based on project
def highlight_project(row):
    if row['Project'] in ['P1', 'P2', 'P3', 'P4', 'P5']:
        return ['background-color: yellow']*len(row)
    else:
        return ['']*len(row)

# create a new dataframe with the highlighted rows
highlighted_df = df.style.apply(highlight_project, axis=1).set_table_styles([{'selector': '', 'props': [('border', '1px solid black')]}]).render()

# convert the Styler object to a Pandas dataframe
highlighted_df = pd.read_html(highlighted_df, header=0, index_col=0)[0]

# write the highlighted dataframe to the same Excel sheet
book = load_workbook('filename.xlsx')
writer = pd.ExcelWriter('filename.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for sheetname in writer.sheets:
    sheet = writer.sheets[sheetname]
    for idx, row in enumerate(dataframe_to_rows(highlighted_df)):
        if idx == 0:
            continue
        sheet.cell(row=idx+1, column=1).value = ''
        for col in range(1, len(row)+1):
            sheet.cell(row=idx+1, column=col+1).value = row[col-1]
writer.save()

print('Done!')
