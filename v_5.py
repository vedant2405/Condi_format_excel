import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Loop through each row in the worksheet
for row in worksheet.iter_rows(min_row=2):
    # Get the project for the current row
    project = row[1].value
    
    # Check if the project is P1 to P5
    if project in ['P1', 'P2', 'P3', 'P4', 'P5']:
        # Highlight the row with yellow background color
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Save the changes to the workbook
workbook.save('example.xlsx')
