import pandas as pd
from openpyxl import load_workbook

# read the Excel sheet into a Pandas dataframe
df = pd.read_excel('filename.xlsx')

# define a function to highlight rows based on project
def highlight_project(row):
    if row['Project'] in ['P1', 'P2', 'P3', 'P4', 'P5']:
        return 'background-color: yellow'
    else:
        return ''

# create a HTML-style string with the highlighted rows
highlighted_rows = df.apply(highlight_project, axis=1).to_list()
html_str = df.to_html(index=False, border=1, classes='table table-striped table-hover', table_id='myTable',
                      header="true", justify="left", col_space="10px", row_styles=highlighted_rows)

# write the HTML string to the same Excel sheet
book = load_workbook('filename.xlsx')
writer = pd.ExcelWriter('filename.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for sheetname in writer.sheets:
    sheet = writer.sheets[sheetname]
    sheet.reset_dimensions()
    for row in pd.read_html(html_str)[0].itertuples():
        for idx, cell_value in enumerate(row[1:], 1):
            cell = sheet.cell(row=row[0]+2, column=idx)
            cell.value = cell_value
    writer.save()

print('Done!')
